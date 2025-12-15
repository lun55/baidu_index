'''
    百度搜索指数爬虫
    by: lun55
'''

import requests
import json
import os
from datetime import datetime, timedelta
import openpyxl
from time import sleep
import time
import random

area_code = { 
    # "909": "福建",
    "50": "福州",
    "51": "莆田",
    "52": "三明",
    "53": "龙岩",
    "54": "厦门",
    "55": "泉州",
    "56": "漳州",
    "87": "宁德",
    "253": "南平",
    }

# 解码函数
def decrypt(ptbk, index_data):
    n = len(ptbk)//2
    a = dict(zip(ptbk[:n], ptbk[n:]))
    return "".join([a[s] for s in index_data])

def request_with_retry(url, headers=None, proxies=None, max_retries=3):
    for attempt in range(max_retries):
        try:
            # 随机延迟（第1次2秒，第2次4秒...）
            delay = 2 * (attempt + 1) + random.uniform(-0.5, 0.5)
            time.sleep(delay)
            
            # 发送请求
            response = requests.get(url, headers=headers, proxies=proxies,timeout=10)
            response.raise_for_status()
            return response.json()
        
        except requests.exceptions.RequestException as e:
            print(f"第 {attempt + 1} 次尝试失败: {e}")
            if attempt == max_retries - 1:
                return None

# 获取数据源并暂存至文件中
def get_index_data(keys,regionCode, year):
    words = [[{"name": keys, "wordType": 1}]]
    words = str(words).replace(" ", "").replace("'", "\"")
    startDate = f"{year}-01-01"
    endDate = f"{year}-12-31"
    url = f'http://index.baidu.com/api/SearchApi/index?area={regionCode}&word={words}&startDate={startDate}&endDate={endDate}'

    proxy_url = "http://api.xiequ.cn/VAD/GetIp.aspx?act=get&uid=160867&vkey=31052E916B5AD93F47A86A61A78849F6&num=1&time=30&plat=0&re=0&type=0&so=1&ow=1&spl=1&addr=&db=1"
    # 请求头配置
    headers = {
        "Connection": "keep-alive",
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Cipher-Text": "1748491566391_1748569584082_rGNNc2tKf78QCtJvr0GDVjRQUkPrpQcDtvTr+MR+wj5ONWxlQbUr69reege9U/RamxSAVilPm1LjlrE8mDdwTCnqpc0HQmLUkXZEjUYy7/TrfaRVzmxEAoowLBWflfomMG6lpYBanu8RvA9PNpU37v2TXnEfNmu+wpIdXIyjKaRRHIWpdmV7JMq/TH1rKpjm7G0/pweze2JUnhcldlHKOESJ5lKt9lijZ9YmuuyPMsTYAS8BeOCaDy21DfgovOsDra9RmzlvgateMKC8OmzLGPKUECRgxLuAMmaz7KZLSUtCuSqoIW0XhBVEqsUq4j1+xPYmvQn0NvkYK4dLQEmZK3y1AOuOBxIJfIWAH1tDf8CXMJ6N/F8GSagV1dm+AZT9T7klMWK6mGWJ8IODIjXL9WVmrDz0JTSxsSe16K9AbmtqVTTbYvZaGr6cD+uGzUTB",
        "Referer": "https://index.baidu.com/v2/main/index.html",
        "Accept-Language": "zh-CN,zh;q=0.9",
        'Cookie': Cookie}
    # 使用代理池
    # response = request_with_retry(proxy_url)
    # ip = response.get('data')[0].get('IP')
    # port = response.get('data')[0].get('Port')
    # if not ip:
    #     proxies = None
    #     print("获取代理失败")
    # else:
    #     proxy_ip = ip + ":" + str(port)
    #     proxies = {
    #     'http': 'http://' + proxy_ip,  # 例如 'http://123.45.67.89:8080'
    #     'https': 'http://' + proxy_ip  # 例如 'https://123.45.67.89:8080'
    #     }
    res_json = request_with_retry(url, headers=headers,proxies=None)
    # print(res_json)
    if not res_json:
        print("抓取关键词："+keys+" 失败")
        return None,None 
    else:
        # 获取特征值
        data = res_json['data']
        # print(data)
        uniqid = data["uniqid"]
        url = f'http://index.baidu.com/Interface/ptbk?uniqid={uniqid}'
        # 获取解码字
        ptbk = request_with_retry(url, headers=headers)['data']

        #创建暂存文件夹
        path = os.path.join('res', area_code[regionCode])
        os.makedirs(path, exist_ok=True)
        filename = f"{keys}_{area_code[regionCode]}_{year}.json"
        file_path = os.path.join(path, filename)
        with open(file_path, 'w', encoding='utf-8') as json_file:
            json.dump(res_json, json_file, ensure_ascii=False, indent=4)
        return file_path, ptbk

def reCode(file_path, ptbk):
    # 读取暂存文件
    with open(file_path, 'r', encoding='utf-8') as file:
        res = json.load(file)
    
    data = res['data']
    user_index = data['userIndexes'][0]
    start_date_str = user_index['all']['startDate']
    end_date_str = user_index['all']['endDate']
    
    # 解析日期范围

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    date_range = (end_date - start_date).days + 1  # 实际数据天数

    result = {}
    name = user_index['word'][0]['name']
    result["name"] = name
    
    # 处理数据
    raw_data = user_index['all']['data']
    if not raw_data:  # 无数据情况
        result["data"] = [0] * date_range
    else:
        try:
            # 解密数据
            decrypted = decrypt(ptbk, raw_data)
            data_points = [int(x) if x != '' else 0 for x in decrypted.split(",")]
            
            # 确保数据长度与日期范围匹配
            if len(data_points) < date_range:
                # 在末尾补0
                data_points.extend([0] * (date_range - len(data_points)))
            elif len(data_points) > date_range:
                # 截断多余数据
                data_points = data_points[:date_range]
                
            result["data"] = data_points
        except Exception as e:
            print(f"数据处理错误: {e}")
            result["data"] = [0] * date_range
    
    return result

#创建日期表格
def create_excel(key, regionCode, start_year, end_year):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 设置第一行的标题
    sheet['A1'] = '日期'

    # 计算日期范围
    start_date = datetime(start_year, 1, 1)
    end_date = datetime(end_year, 12, 31)
    current_date = datetime.now()  # 获取当前日期

    # 如果 end_date 超过当前日期，则设为当前日期
    if end_date > current_date:
        end_date = current_date
    # 逐行填充日期
    current_date = start_date
    row = 2  # 从第二行开始
    while current_date <= end_date:
        sheet[f'A{row}'] = current_date.strftime('%Y-%m-%d')
        current_date += timedelta(days=1)
        row += 1

    # 保存 Excel 文件
    filename = f'百度指数数据-{key}-{area_code[regionCode]}-{start_year}-{end_year}.xlsx'
    if os.path.exists(filename):
        print(f"文件 {filename} 已存在，跳过写入")
        return None
    workbook.save(filename)
    return filename

#为文件写入数据
def write_to_excel(file_name, name, data, i):
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_name)
        # 获取默认的工作表（第一个工作表）
        sheet = workbook.active
        # 将名称写入第一行第i列
        sheet.cell(row=1, column=i, value=name)
        # 将数据写入从第二行开始的第i列
        for index, value in enumerate(data, start=2):
            sheet.cell(row=index, column=i, value=value)
        # 保存文件
        workbook.save(file_name)
        if len(data) != 0 :
            print(f"关键词-{name}-写入成功!有效数据共{len(data)}个")

    except Exception as e:
        print(f"发生错误: {e}")


def main(keys,regionCode,startDate,endDate):

    data = []
    i = 2
    for key in keys:
        filename = create_excel(key, regionCode,startDate,endDate)
        if not filename:
            continue
        print(filename +"--创建成功！")
        for year in range(startDate, endDate + 1):
            print(f"正在处理第{year}年，请耐心等待……")
            file_path,ptbk = get_index_data(key,regionCode, year)
            if not file_path:
                os.remove(filename)
                print(f"删除文件{filename}")
                sleep(random.uniform(25.5, 35.5))
                break
            res = reCode(file_path,ptbk)
            name = res["name"]
            temp = res["data"]
            data = data + temp
            sleep(random.uniform(2.5, 3.5))  # 3秒 ± 0.5秒波动
        # print(data)
        write_to_excel(filename,name,data,i)
        data = []
    print("程序运行结束！")


if __name__ == '__main__':

    area_codes = area_code.keys()
    # 参数列表
    Cookie = 'BAIDUID=8860F3AFC8172EC384D2AAC6808DEEB2:FG=1; BAIDUID_BFESS=8860F3AFC8172EC384D2AAC6808DEEB2:FG=1; H_WISE_SIDS=110085_287279_626068_628198_632160_633611_642955_644661_645092_645169_645922_646538_646561_647614_647658_648250_645030_648502_648445_648439_648717_648996_649054_649073_649060_649047_648586_649037_649326_646542_649351_649343_649649_649753_649868_649776_649910_649957_650069_650054_649885_650086_650257_650288_650285_650418_650521_648982_650745_650013_650760_650797_645236_649494_650827_651003_651000_650992_650995_650151_650903_651060_651074_651177_651185_651273_650556_651292_651312_649929_651371_651378_651076_651391_651418_651420_651416_647692_651504_651514_651531_651536_651537_651553_651574_651559_651555_651565_651580_651594_651606_651601_651587_651582_651608_651613_641261_651485_651486_651489; H_WISE_SIDS_BFESS=110085_287279_626068_628198_632160_633611_642955_644661_645092_645169_645922_646538_646561_647614_647658_648250_645030_648502_648445_648439_648717_648996_649054_649073_649060_649047_648586_649037_649326_646542_649351_649343_649649_649753_649868_649776_649910_649957_650069_650054_649885_650086_650257_650288_650285_650418_650521_648982_650745_650013_650760_650797_645236_649494_650827_651003_651000_650992_650995_650151_650903_651060_651074_651177_651185_651273_650556_651292_651312_649929_651371_651378_651076_651391_651418_651420_651416_647692_651504_651514_651531_651536_651537_651553_651574_651559_651555_651565_651580_651594_651606_651601_651587_651582_651608_651613_641261_651485_651486_651489; ZFY=oO:BMS7WBYVKeOw1bY7e:BSOli:Av2XsX58MWfKmmlIK4Y:C; Hm_lvt_d101ea4d2a5c67dab98251f0b5de24dc=1748433299,1748568191,1748937942,1748955545; HMACCOUNT=63D1138021DFD966; ppfuid=FOCoIC3q5fKa8fgJnwzbE67EJ49BGJeplOzf+4l4EOvDuu2RXBRv6R3A1AZMa49I27C0gDDLrJyxcIIeAeEhD8JYsoLTpBiaCXhLqvzbzmvy3SeAW17tKgNq/Xx+RgOdb8TWCFe62MVrDTY6lMf2GrfqL8c87KLF2qFER3obJGm51EODDlnqgz44AdUN5VVLGEimjy3MrXEpSuItnI4KD321Mdgb7z54NfprT7nU1gfI5V0fRa0WHd32lXFKuen2eOqwkUSI0R4ks430johpgRJsVwXkGdF24AsEQ3K5XBbh9EHAWDOg2T1ejpq0s2eFy9ar/j566XqWDobGoNNfmfpaEhZpob9le2b5QIEdiQcF+6iOKqU/r67N8lf+wxW6FCMUN0p4SXVVUMsKNJv2T2Q0Rs14gDuqHJ3rxHJuOGO4LkPV+7TROLMG0V6r0A++zkWOdjFiy1eD/0R8HcRWYo64YZQejZKa7nFsdjKdPqCp+HBavJhpxl858h16cMtKQmxzisHOxsE/KMoDNYYE7ucLE22Bi0Ojbor7y6SXfVj7+B4iuZO+f7FUDWABtt/WWQqHKVfXMaw5WUmKnfSR5wwQa+N01amx6X+p+x97kkGmoNOSwxWgGvuezNFuiJQdt51yrWaL9Re9fZveXFsIu/gzGjL50VLcWv2NICayyI8BE9m62pdBPySuv4pVqQ9Sl1uTC//wIcO7QL9nm+0N6JgtCkSAWOZCh7Lr0XP6QztjlyD3bkwYJ4FTiNanaDaDCdfMiwHbmJJlyOwfxZhsS2GLXPRZALgEnCalLVRvIh54ddVfMNQzGka31pGZdK0uSGk66HDxtjKMU4HPNa0dthF7UsHf7NW9eE+gwuTQSa7GLWfOy9+ap4iFBQsmjpefgOF89jAHLbnVUejtrqqvdVSQ/4gzJOb0DGzeEZ5GeyPSEfcJNw4xf285NSi5WN28XlV3f0HZXSpuSxTnDK9hXC355Jsoo1+Hl0sO4qqegfHgBRkDPTwtXjYtgzmW74m0fDU2MZaxpBZZF8YurfocYcmDdcxFKeoIFQmVqAoAU+3YcXQt2xKThZZyV1v3sCvnzidUZtKM9cRRUfRWBtQSb50APM+gs/408xg7KHCB8AOKpZpfIpPhQ0RJhew8GR0aTqYsJo1IRCwM3UbbrvtJ7eqPMNzJcGcSYcQWm1FubInMonve94c+p8Vi2wc72MfReeFiTzMp1G6pDt2e40gPDGbdQI+jba4UjRlyA+9CbTXeo25gLlwEeBBlbK/gwjLfyky6FPenbJgt/vQK9TAiTA==; BDUSS=F3bWdMdDNxNXRrck5TUVJoZ1BFV1hKWXNKSEN4OWZzRk1mR2xZNHNna3VKMmRvRVFBQUFBJCQAAAAAAAAAAAEAAAAJJVJaw8662zIzNQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC6aP2gumj9oc3; SIGNIN_UC=70a2711cf1d3d9b1a82d2f87d633bd8a049901115117BKkjhRHpAxCn1ZYfI4xs6ugcnHF3COpzubgTHgPFDZ9VbMalmtbBdTp6wQtJits11tbU9mjY0sCfwobUfb9K7F5Z1Bl8p%2FVSob%2BxWsyqPxvQ5N82kDaZLEd8P6aCdH6aIabB50Pg8W2qwHJecfm2OHgaCn4UQCrXCwFrzATAyOeotrFL4gPD%2BXjrqUVfctC%2F8N9BT8AMNiXgK27g8Tbg1k2b5%2BvDgxXIFckx05lkZQA5BSxqodrfbnH9Qm5piD11D%2FkL02U28cV6jeqekAOqz%2BTOOcj4i3zjU177udNHcI%3D71925260480403258741291123746176; __cas__rn__=499011151; __cas__st__212=bef9322246a05c3f72125937ca8a7c888fd89f16fe10a6ce48085428583d68fe1004d3d9eabbfe5e02444189; __cas__id__212=68652865; CPTK_212=1137210318; CPID_212=68652865; bdindexid=iilrdqmtnrjqbljr674hi9lde1; RT="z=1&dm=baidu.com&si=6622404d-af4b-4a60-bc57-c29243172d76&ss=mbh8ncrh&sl=b&tt=6f6&bcn=https%3A%2F%2Ffclog.baidu.com%2Flog%2Fweirwood%3Ftype%3Dperf"; Hm_lpvt_d101ea4d2a5c67dab98251f0b5de24dc=1748998719; ab_sr=1.0.1_Y2JlNWExNjI5Y2MwMWZhNjhlMjVjOTkxYzBkZTcwMmMyMTExYzYyZWJmYzM2MmEzOWY4MTRmYzRhNDViNmU1MmM4MWZmMGMxOGRjMDgyYWU0OTQzOGQ3MjI1YTE1YWQyNzQ2ZjQ4MjU4ODRiZTM5Yzk3ODJhZTZjYjUzMDM1OGJlZWQxMDVjZjVmNDk4NTBiMWEzY2ViZThlYWQ3NGM3Yw==; BDUSS_BFESS=F3bWdMdDNxNXRrck5TUVJoZ1BFV1hKWXNKSEN4OWZzRk1mR2xZNHNna3VKMmRvRVFBQUFBJCQAAAAAAAAAAAEAAAAJJVJaw8662zIzNQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC6aP2gumj9oc3'
    # 获取的时间区间，若只获取某一年份，则二者相同
    # 注意！年份区间下限为2011年，不建议选择太早年份
    startDate = 2022
    endDate = 2024
    # keys = ['华为云','阿里云','腾讯云','智慧农业','智能制造','智能交通','智慧教育','智慧医疗','智能工厂','数字化转型','智慧园区','工业互联网','数据资产',''
            # '数据资源','数据安全','数据治理']
    keys = ['京东','淘宝','拼多多','亚马逊','朴朴','唯品会','抖音','快手','美团','饿了么']
    # 要搜索的关键词，可以输入一个列表
    for code in area_codes:
        # keys = [area_code[code] + "市政府"]
        if Cookie == "":
            Cookie = input("请输入你的Cookie，若错误则无法运行：")
        elif startDate < 2011:
            print("请注意初始年份限制！！！")
        else:
            main(keys,code,startDate, endDate)

